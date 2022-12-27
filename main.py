import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
import numpy as np

class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.salary = Salary(salary_from, salary_to, False, salary_currency)
        self.area_name = area_name
        self.published_at = published_at

class Salary:
    currency_rub = dict(
        AZN=35.68,
        BYR=23.91,
        EUR=59.90,
        GEL=21.74,
        KGS=0.76,
        KZT=0.13,
        RUR=1,
        UAH=1.64,
        USD=60.66,
        UZS=0.0055,
    )

    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        self.salary_from = int(float(salary_from))
        self.salary_to = int(float(salary_to))
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def convert_to_rub(self):
        return (self.salary_from + self.salary_to) / 2 * self.currency_rub[self.salary_currency]

class Input:
    def input(self):
        fName = input('Введите название файла: ')

        job = input('Введите название профессии: ')

        salary_rub = self.get_dict()
        salary_count = self.get_dict()

        job_rub = self.get_dict()
        job_count = self.get_dict()

        data_objects = []

        with open(fName, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = []

            is_first = True
            for row in reader:
                if is_first:
                    is_first = False
                    header = row
                else:
                    if not "" in row and len(row) == len(header):

                        obj = Vacancy(
                            row[header.index('name')],
                            row[header.index('salary_from')],
                            row[header.index('salary_to')],
                            row[header.index('salary_currency')],
                            row[header.index('area_name')],
                            row[header.index('published_at')]
                        )
                        data_objects.append(obj)

                        year = int(obj.published_at[:4])
                        salary_rub[year] = self.get_medium(salary_rub[year], obj.salary.convert_to_rub(),
                                                           salary_count[year])
                        salary_count[year] += 1

                        if (obj.name.find(job) != -1):
                            job_rub[year] = self.get_medium(job_rub[year], obj.salary.convert_to_rub(), job_count[year])
                            job_count[year] += 1

        salary_rub = self.del_empty(self.round_elements(salary_rub))
        job_rub = self.del_empty(self.round_elements(job_rub))
        salary_count = self.del_empty(salary_count)
        job_count = self.del_empty(job_count)

        print('Динамика уровня зарплат по годам:', salary_rub)
        print('Динамика количества вакансий по годам:', salary_count)
        print('Динамика уровня зарплат по годам для выбранной профессии:', job_rub)
        print('Динамика количества вакансий по годам для выбранной профессии:', job_count)

        city_salary = {}
        city_count = {}
        city_frac = {}

        for it in data_objects:
            city = it.area_name
            if city not in city_salary.keys():
                if len([x for x in data_objects if x.area_name == city]) >= int(len(data_objects) / 100):
                    city_salary[city] = it.salary.convert_to_rub()
                    city_count[city] = 1
            else:
                city_salary[city] = self.get_medium(city_salary[city], it.salary.convert_to_rub(), city_count[city])
                city_count[city] += 1

        all = len(data_objects)
        for key, value in city_count.items():
            city_frac[key] = round(value / (all / 100) / 100, 4)

        city_salary = self.round_elements(self.del_empty(self.sort_city(city_salary)))
        city_frac = self.del_empty(self.sort_city(city_frac))

        print('Уровень зарплат по городам (в порядке убывания):', city_salary)
        print('Доля вакансий по городам (в порядке убывания):', city_frac)

        return job, salary_rub, salary_count, job_rub, job_count, city_salary, city_frac

    def get_dict(self):
        return {x: 0 for x in range(2007, 2023)}

    def del_empty(self, d):
        cd = dict(filter(lambda x: x[1], d.items()))
        if len(cd.keys()) == 0:
            cd[2022] = 0
        return cd

    def get_medium(self, m, x, n):
        return (m * n + x) / (n + 1)

    def sort_city(self, d):
        return dict(sorted(d.items(), key=lambda x: x[1], reverse=True)[:10])

    def round_elements(self, d):
        return dict(map(lambda x: (x[0], int(x[1])), d.items()))

class Report:
    fHeaders = [
        'Год',
        'Средняя зарплата',
        'Средняя зарплата - ',
        'Количество вакансий',
        'Количество вакансий - '
    ]

    sHeaders = [
        'Город',
        'Уровень зарплат',
        '',
        'Город',
        'Доля вакансий'
    ]
    
    def __init__(self):
        self.wb = Workbook()

        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)

        self.wb.create_sheet('Статистика по годам')
        self.wb.create_sheet('Статистика по городам')

    def about_text(self, value):
        if value is None:
            return ""
        return str(value)

    def make_outline(self):
        for row in self.wb.active.rows:
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    def set_size(self):
        for column_cells in self.wb.active.columns:
            length = max(len(self.about_text(cell.value)) for cell in column_cells)
            self.wb.active.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    def create_first_stat(self, data):
        self.wb.active = self.wb['Статистика по годам']
        ws = self.wb.active

        self.fHeaders[2] = self.fHeaders[2] + data[0]
        self.fHeaders[4] = self.fHeaders[4] + data[0]
        ws.append(it for it in self.fHeaders)
        for row in ws.rows:
            for cell in row:
                cell.font = Font(bold=True)

        for year in data[1].keys():
            row = [year, data[1][year], data[3][year], data[2][year], data[4][year]]
            ws.append(row)

        self.set_size()
        self.make_outline()


    def create_second_stat(self, data):
        self.wb.active = self.wb['Статистика по городам']
        ws = self.wb.active

        ws.append(it for it in self.sHeaders)
        for row in ws.rows:
            for cell in row:
                cell.font = Font(bold=True)

        info1 = list(data[0].keys())
        info2 = list(data[0].values())
        info3 = list(data[1].keys())
        info4 = list(data[1].values())

        for i in range(len(data[0])):
            row = [info1[i], info2[i], '', info3[i], info4[i]]
            ws.append(row)

        self.set_size()
        self.make_outline()

        for i in range(1, 12):
            ws[f"C{i}"].border = Border()

        for i in range(1, 12):
            ws[f"E{i}"].number_format = '0.00%'

        self.wb.active = self.wb['Статистика по годам']

    def generate_excel(self, data1, data2):
        self.create_first_stat(data1)
        self.create_second_stat(data2)

        self.wb.save('report.xlsx')

    def create_ysalary(self, job, data1, data2, ax):
        labels = list(data1.keys())
        average = list(data1.values())
        jobs = list(data2.values())

        x = np.arange(len(labels))
        width = 0.35

        ax.bar(x - width / 2, average, width, label='средняя з/п')
        ax.bar(x + width / 2, jobs, width, label=f"з/п {job}")

        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(prop={"size": 8})
        ax.grid(axis='y')
        ax.tick_params(axis='both', labelsize=8)

    def create_cjobs(self, data, ax):
        x = list(data.values())
        x.append(1 - sum(x))
        cities = list(data.keys()) + ['Другие']

        ax.set_title('Доля вакансий по городам')
        ax.pie(x, labels=cities, textprops={'fontsize': 6}, startangle=90)

    def create_ycounts(self, job, data1, data2, ax):
        labels = list(data1.keys())
        counts = list(data1.values())
        jobs = list(data2.values())

        x = np.arange(len(labels))
        width = 0.35

        ax.bar(x - width / 2, counts, width, label='Количество вакансий')
        ax.bar(x + width / 2, jobs, width, label=f"Количество вакансий\n{job}")

        ax.set_title('Количество вакансий по годам')
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(prop={"size": 8})
        ax.grid(axis='y')
        ax.tick_params(axis='both', labelsize=8)

    def create_csalary(self, data, ax):
        sep = lambda x: x.replace(' ', '\n').replace('-', '\n')

        cities = list(map(sep, data.keys()))[::-1]
        values = list(data.values())[::-1]
        y_pos = np.arange(len(cities))

        ax.barh(y_pos, values)
        ax.set_yticks(y_pos, labels=cities, fontsize=6)
        ax.set_title('Уровень зарплат по городам')
        ax.tick_params(axis='x', labelsize=8)

    def get_png(self, data):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        self.create_ycounts(data[0], data[2], data[4], ax2)
        self.create_csalary(data[5], ax3)
        self.create_ysalary(data[0], data[1], data[3], ax1)
        self.create_cjobs(data[6], ax4)

        fig.tight_layout()
        fig.savefig('graph.png')


ic = Input()
data = ic.input()

report = Report()
report.generate_excel(data[:5], data[5:])
report.get_png(data)
